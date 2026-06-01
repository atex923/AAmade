import json
import re
import ast
import operator
import tkinter as tk
from datetime import datetime
from pathlib import Path
from tkinter import filedialog, messagebox
from tkinter import ttk


CODE_ORDER = "123456789ABCDEFGHIJKLMNOPQRSTUVWXYZ"
CODE_PATTERN = re.compile(r"^0[1-9A-Z]{0,7}$")
DATA_COLUMNS = 10
MIN_DATA_ROWS = 20
ITEM_COLUMN_INDEX = 0
CODE_COLUMN_INDEX = 1
SAVE_EXTENSION = ".json"
APP_NAME = "AA預算編輯器"
APP_VERSION = "V0.1.j"
DEFAULT_COVER_TEMPLATE = Path("/Users/atex1/Desktop/預算書封面.xlsx")
QTY_COLUMNS = 6
QTY_EXPR_INDEX = 4
QTY_RESULT_INDEX = 5
QTY_SOURCE_COLUMNS = (0, 1, 2, 4, None, 6)
QTY_SYNC_COLUMNS = (0, 1, 2, 3, 5)
QTY_EDITABLE_COLUMNS = (2, 3, 4)
QTY_OPERATORS = {
    ast.Add: operator.add,
    ast.Sub: operator.sub,
    ast.Mult: operator.mul,
    ast.Div: operator.truediv,
    ast.USub: operator.neg,
    ast.UAdd: operator.pos,
}

COVER_TEXT_FIELDS = [
    ("project_name", "工程名稱", "B2"),
    ("budget_year", "預算年度", "B3"),
    ("location", "工程地點", "F3"),
    ("account_subject", "會計科目", "B4"),
    ("start_date", "預定開工日期", "F4"),
    ("project_no", "工程編號", "B5"),
    ("finish_date", "預定竣工日期", "F5"),
    ("summary", "工程概要", "B6"),
    ("note_tax", "附註一", "H8"),
    ("note_days", "預定工期", "H9"),
    ("note_total_label", "總預算標籤", "H17"),
    ("note_total", "總預算", "H18"),
]

COVER_AMOUNT_ROWS = [
    ("contract", "發包工程費", 8),
    ("self_labor", "自辦工費", 9),
    ("self_material", "自購材料費", 10),
    ("rail_material", "路購材料費", 11),
    ("reserved_material", "路備材料費", 12),
    ("management", "管理費", 13),
    ("agency", "代辦服務費", 14),
    ("freight", "運雜費", 15),
    ("other", "其他", 16),
    ("tax", "營業稅", 17),
    ("total", "合計", 18),
]

COVER_ATTACHMENTS = [
    ("construction_budget_pages", "施工預算明細表頁數", "C19"),
    ("unit_analysis_pages", "工程單價分析表頁數", "C20"),
    ("quantity_calc_pages", "工程數量計算表頁數", "C21"),
    ("engineering_note_pages", "工程說明書頁數", "F19"),
]


class BudgetEditor(tk.Tk):
    def __init__(self):
        super().__init__()

        self.title(f"{APP_NAME} {APP_VERSION}")
        self.geometry("1180x720")
        self.minsize(980, 560)

        self.rows_data = [["" for _column in range(DATA_COLUMNS)] for _row in range(MIN_DATA_ROWS)]
        self.quantity_data = []
        self.cell_vars = []
        self.cell_entries = []
        self.quantity_vars = []
        self.quantity_entries = []
        self.quantity_col_widths = [100, 160, 220, 100, 220, 140]
        self.quantity_resize = None
        self.level_labels = []
        self.code_entries = []
        self.code_status_labels = []
        self.row_widgets = []
        self.table_parent = None
        self.table_canvas = None
        self.cover_canvas = None
        self.quantity_parent = None
        self.quantity_canvas = None
        self.main_notebook = None
        self.focused_cell = (0, CODE_COLUMN_INDEX)
        self.status_var = tk.StringVar(value="請從第 2 列開始輸入資料。")
        self.hidden_level = None
        self.focus_code = ""
        self.deleted_rows_stack = []
        self.undo_stack = []
        self.redo_stack = []
        self._history_suspended = False
        self._edit_snapshot = None
        self.editing_cell = None
        self._is_validating = False
        self._is_rendering = False

        self._build_layout()

    def _build_layout(self):
        self.columnconfigure(0, weight=1)
        self.rowconfigure(0, weight=0)
        self.rowconfigure(1, weight=1)

        top_area = ttk.Frame(self, padding=12)
        top_area.grid(row=0, column=0, sticky="nsew")
        top_area.columnconfigure(1, weight=3)
        top_area.columnconfigure(3, weight=1)
        top_area.columnconfigure(5, weight=1)

        ttk.Label(top_area, text="工程名稱：").grid(row=0, column=0, sticky="w", padx=(0, 6))
        self.project_name_var = tk.StringVar()
        ttk.Entry(top_area, textvariable=self.project_name_var).grid(row=0, column=1, sticky="ew", padx=(0, 12))

        ttk.Label(top_area, text="執行號：").grid(row=0, column=2, sticky="w", padx=(0, 6))
        self.execution_no_var = tk.StringVar()
        ttk.Entry(top_area, textvariable=self.execution_no_var, width=18).grid(
            row=0, column=3, sticky="ew", padx=(0, 12)
        )

        ttk.Label(top_area, text="預算年度").grid(row=0, column=4, sticky="w", padx=(0, 6))
        self.budget_year_var = tk.StringVar()
        ttk.Entry(top_area, textvariable=self.budget_year_var, width=12).grid(row=0, column=5, sticky="ew")

        ttk.Label(top_area, text="工程位置：").grid(row=1, column=0, sticky="w", padx=(0, 6), pady=(8, 0))
        self.project_location_var = tk.StringVar()
        ttk.Entry(top_area, textvariable=self.project_location_var).grid(
            row=1, column=1, columnspan=5, sticky="ew", pady=(8, 0)
        )

        ttk.Label(top_area, text="工程內容：").grid(row=2, column=0, sticky="nw", padx=(0, 6), pady=(8, 0))
        self.project_content_text = tk.Text(top_area, height=4, wrap="word", relief="solid", borderwidth=1)
        self.project_content_text.grid(row=2, column=1, columnspan=5, sticky="ew", pady=(8, 0))

        tools = ttk.Frame(top_area)
        tools.grid(row=3, column=0, columnspan=6, sticky="ew", pady=(10, 0))
        tools.columnconfigure(12, weight=1)

        ttk.Label(tools, text="插入行位數量").grid(row=0, column=0, sticky="w", padx=(0, 6))
        self.insert_count_var = tk.StringVar(value="1")
        insert_count = ttk.Spinbox(
            tools,
            from_=1,
            to=100,
            textvariable=self.insert_count_var,
            width=6,
            justify="center",
        )
        insert_count.grid(row=0, column=1, sticky="w", padx=(0, 8))

        insert_button = ttk.Button(tools, text="插入行位", command=self.insert_rows_at_cursor)
        insert_button.grid(row=0, column=2, sticky="w", padx=(0, 12))

        delete_button = ttk.Button(tools, text="刪除行位", command=self.delete_row_at_cursor)
        delete_button.grid(row=0, column=3, sticky="w", padx=(0, 16))

        ttk.Button(tools, text="復原刪除", command=self.undo_delete_row).grid(row=0, column=4, sticky="w", padx=(0, 16))
        ttk.Button(tools, text="回復修改", command=self.undo_last_change).grid(row=0, column=13, sticky="w", padx=(0, 8))
        ttk.Button(tools, text="重做修改", command=self.redo_last_change).grid(row=0, column=14, sticky="w", padx=(0, 8))

        ttk.Label(tools, text="隱藏階層").grid(row=0, column=5, sticky="w", padx=(0, 6))
        self.hide_level_var = tk.StringVar()
        hide_level = ttk.Spinbox(
            tools,
            from_=1,
            to=8,
            textvariable=self.hide_level_var,
            width=5,
            justify="center",
        )
        hide_level.grid(row=0, column=6, sticky="w", padx=(0, 8))
        ttk.Button(tools, text="套用", command=self.hide_level_rows).grid(row=0, column=7, sticky="w", padx=(0, 16))

        ttk.Label(tools, text="指定工項").grid(row=0, column=8, sticky="w", padx=(0, 6))
        self.show_code_var = tk.StringVar()
        show_code = ttk.Entry(tools, textvariable=self.show_code_var, width=16)
        show_code.grid(row=0, column=9, sticky="w", padx=(0, 8))
        ttk.Button(tools, text="顯示以下階層", command=self.show_only_code_tree).grid(row=0, column=10, sticky="w", padx=(0, 12))
        ttk.Button(tools, text="恢復顯示全部", command=self.show_all_rows).grid(row=0, column=11, sticky="w", padx=(0, 16))

        self.top_message_var = tk.StringVar(value="插入位置會從目前游標所在列的下一行開始插入。")

        ttk.Label(tools, text="跳到工項").grid(row=1, column=0, sticky="w", padx=(0, 6), pady=(8, 0))
        self.jump_code_var = tk.StringVar()
        ttk.Entry(tools, textvariable=self.jump_code_var, width=16).grid(row=1, column=1, sticky="w", padx=(0, 8), pady=(8, 0))
        ttk.Button(tools, text="跳到", command=self.jump_to_code).grid(row=1, column=2, sticky="w", padx=(0, 12), pady=(8, 0))

        ttk.Label(tools, text="搜尋關鍵字").grid(row=1, column=3, sticky="w", padx=(0, 6), pady=(8, 0))
        self.search_keyword_var = tk.StringVar()
        ttk.Entry(tools, textvariable=self.search_keyword_var, width=18).grid(row=1, column=4, sticky="w", padx=(0, 8), pady=(8, 0))
        ttk.Button(tools, text="搜尋", command=self.search_keyword).grid(row=1, column=5, sticky="w", padx=(0, 16), pady=(8, 0))

        ttk.Label(tools, text="檔名").grid(row=1, column=6, sticky="w", padx=(0, 6), pady=(8, 0))
        self.save_filename_var = tk.StringVar()
        ttk.Entry(tools, textvariable=self.save_filename_var, width=20).grid(row=1, column=7, sticky="w", padx=(0, 8), pady=(8, 0))
        ttk.Button(tools, text="儲存記錄", command=self.save_record).grid(row=1, column=8, sticky="w", padx=(0, 8), pady=(8, 0))
        ttk.Button(tools, text="開啟記錄", command=self.open_record).grid(row=1, column=9, sticky="w", padx=(0, 8), pady=(8, 0))
        ttk.Button(tools, text="輸出第2分頁", command=self.export_cover_content_excel).grid(
            row=1, column=10, sticky="w", padx=(0, 8), pady=(8, 0)
        )
        ttk.Button(tools, text="執行第3分頁計算", command=self.calculate_quantity_page).grid(
            row=1, column=11, sticky="w", padx=(0, 8), pady=(8, 0)
        )

        lower_area = ttk.Frame(self, padding=(12, 4, 12, 12))
        lower_area.grid(row=1, column=0, sticky="nsew")
        lower_area.columnconfigure(0, weight=1)
        lower_area.rowconfigure(0, weight=1)

        self.main_notebook = ttk.Notebook(lower_area)
        self.main_notebook.grid(row=0, column=0, sticky="nsew")

        table_shell = ttk.Frame(self.main_notebook)
        self.main_notebook.add(table_shell, text="第1頁 試算表")
        table_shell.columnconfigure(0, weight=1)
        table_shell.rowconfigure(0, weight=1)

        canvas = tk.Canvas(table_shell, highlightthickness=0)
        self.table_canvas = canvas
        y_scroll = ttk.Scrollbar(table_shell, orient="vertical", command=canvas.yview)
        x_scroll = ttk.Scrollbar(table_shell, orient="horizontal", command=canvas.xview)
        canvas.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)

        canvas.grid(row=0, column=0, sticky="nsew")
        y_scroll.grid(row=0, column=1, sticky="ns")
        x_scroll.grid(row=1, column=0, sticky="ew")

        self.table_parent = ttk.Frame(canvas)
        canvas_window = canvas.create_window((0, 0), window=self.table_parent, anchor="nw")

        self.table_parent.bind(
            "<Configure>",
            lambda event: canvas.configure(scrollregion=canvas.bbox("all")),
        )
        canvas.bind(
            "<Configure>",
            lambda event: canvas.itemconfigure(canvas_window, width=max(event.width, 1120)),
        )
        canvas.bind_all("<MouseWheel>", self._on_mouse_wheel)
        canvas.bind_all("<Button-4>", self._on_mouse_wheel)
        canvas.bind_all("<Button-5>", self._on_mouse_wheel)

        cover_shell = ttk.Frame(self.main_notebook)
        self.main_notebook.add(cover_shell, text="第2頁 預算書封面內容")
        cover_shell.columnconfigure(0, weight=1)
        cover_shell.rowconfigure(0, weight=1)
        self._build_cover_content_page(cover_shell)

        quantity_shell = ttk.Frame(self.main_notebook)
        self.main_notebook.add(quantity_shell, text="第3頁 工項數量計算表")
        quantity_shell.columnconfigure(0, weight=1)
        quantity_shell.rowconfigure(0, weight=1)
        self._build_quantity_page(quantity_shell)

        status = ttk.Label(self, textvariable=self.status_var, anchor="w", padding=(12, 0, 12, 8))
        status.grid(row=2, column=0, sticky="ew")
        message = ttk.Label(self, textvariable=self.top_message_var, anchor="w", foreground="#666666", padding=(12, 0, 12, 8))
        message.grid(row=3, column=0, sticky="ew")

        self._build_table()

    def _build_cover_content_page(self, parent):
        canvas = tk.Canvas(parent, highlightthickness=0)
        self.cover_canvas = canvas
        y_scroll = ttk.Scrollbar(parent, orient="vertical", command=canvas.yview)
        x_scroll = ttk.Scrollbar(parent, orient="horizontal", command=canvas.xview)
        canvas.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)

        canvas.grid(row=0, column=0, sticky="nsew")
        y_scroll.grid(row=0, column=1, sticky="ns")
        x_scroll.grid(row=1, column=0, sticky="ew")

        content = ttk.Frame(canvas)
        window = canvas.create_window((0, 0), window=content, anchor="nw")
        content.bind("<Configure>", lambda _event: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>", lambda event: canvas.itemconfigure(window, width=max(event.width, 980)))

        if DEFAULT_COVER_TEMPLATE.exists():
            self._build_cover_template_grid(content)
        else:
            self._build_cover_content_list(content)

    def _build_cover_template_grid(self, parent):
        try:
            from openpyxl import load_workbook
            from openpyxl.utils import get_column_letter
        except ImportError:
            self._build_cover_content_list(parent)
            return

        workbook = load_workbook(DEFAULT_COVER_TEMPLATE, data_only=False)
        sheet = workbook.active
        merged_map = {}
        skip_cells = set()
        for merged_range in sheet.merged_cells.ranges:
            min_col, min_row, max_col, max_row = merged_range.bounds
            if max_col < 1 or min_col > 15 or max_row < 1 or min_row > 21:
                continue

            min_col = max(min_col, 1)
            max_col = min(max_col, 15)
            min_row = max(min_row, 1)
            max_row = min(max_row, 21)
            merged_map[(min_row, min_col)] = (max_row - min_row + 1, max_col - min_col + 1)
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    if (row, col) != (min_row, min_col):
                        skip_cells.add((row, col))

        for col in range(1, 16):
            letter = get_column_letter(col)
            width = sheet.column_dimensions[letter].width or 8.43
            parent.columnconfigure(col - 1, minsize=max(24, int(width * 8)), weight=0)

        for row in range(1, 22):
            height = sheet.row_dimensions[row].height or 18
            parent.rowconfigure(row - 1, minsize=max(20, int(height * 1.35)), weight=0)

        for row in range(1, 22):
            for col in range(1, 16):
                if (row, col) in skip_cells:
                    continue

                cell = sheet.cell(row=row, column=col)
                rowspan, colspan = merged_map.get((row, col), (1, 1))
                label = self._template_cell_label(parent, cell)
                label.grid(
                    row=row - 1,
                    column=col - 1,
                    rowspan=rowspan,
                    columnspan=colspan,
                    sticky="nsew",
                )

    def _build_cover_content_list(self, content):
        title = tk.Label(
            content,
            text="預算書封面內容",
            bg="#1f4e78",
            fg="#ffffff",
            font=("Arial", 16, "bold"),
            relief="solid",
            borderwidth=1,
            padx=8,
            pady=8,
        )
        title.grid(row=0, column=0, columnspan=6, sticky="nsew")

        note = tk.Label(
            content,
            text="資料來源與對應欄位先保留待設定，本頁只顯示封面內容項目。",
            bg="#ffffff",
            fg="#666666",
            anchor="w",
            relief="solid",
            borderwidth=1,
            padx=8,
            pady=6,
        )
        note.grid(row=1, column=0, columnspan=6, sticky="nsew")

        headers = ["類別", "欄位代號", "封面內容", "封面位置", "資料來源", "對應欄位"]
        widths = [18, 24, 24, 14, 24, 24]
        for column_index, header in enumerate(headers):
            label = self._grid_label(content, header, "#d9eaf7", bold=True, width=widths[column_index])
            label.grid(row=2, column=column_index, sticky="nsew")
            content.columnconfigure(column_index, weight=1, minsize=widths[column_index] * 9)

        for row_index, row_values in enumerate(self._cover_content_rows(), start=3):
            for column_index, value in enumerate(row_values):
                bg = "#e2f0d9" if column_index == 0 else "#ffffff"
                label = self._grid_label(content, value, bg, width=widths[column_index], anchor="w")
                label.grid(row=row_index, column=column_index, sticky="nsew")

    def _cover_content_rows(self):
        rows = []
        rows.extend(("基本資料", key, label, cell, "", "") for key, label, cell in COVER_TEXT_FIELDS)
        rows.extend(("費用明細-預算金額", f"{key}_budget", label, f"B{row}", "", "") for key, label, row in COVER_AMOUNT_ROWS)
        rows.extend(("費用明細-審核金額", f"{key}_audit", label, f"E{row}", "", "") for key, label, row in COVER_AMOUNT_ROWS)
        rows.extend(("附件頁數", key, label, cell, "", "") for key, label, cell in COVER_ATTACHMENTS)
        return rows

    def _build_quantity_page(self, parent):
        canvas = tk.Canvas(parent, highlightthickness=0)
        self.quantity_canvas = canvas
        y_scroll = ttk.Scrollbar(parent, orient="vertical", command=canvas.yview)
        x_scroll = ttk.Scrollbar(parent, orient="horizontal", command=canvas.xview)
        canvas.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)

        canvas.grid(row=0, column=0, sticky="nsew")
        y_scroll.grid(row=0, column=1, sticky="ns")
        x_scroll.grid(row=1, column=0, sticky="ew")

        self.quantity_parent = ttk.Frame(canvas)
        window = canvas.create_window((0, 0), window=self.quantity_parent, anchor="nw")
        self.quantity_parent.bind("<Configure>", lambda _event: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>", lambda event: canvas.itemconfigure(window, width=max(event.width, 900)))
        self._build_quantity_table()

    def _build_quantity_table(self):
        if self.quantity_parent is None:
            return

        self._sync_quantity_source()
        for child in self.quantity_parent.winfo_children():
            child.destroy()

        self.quantity_vars = []
        self.quantity_entries = []
        headers = ["項次", "工項號碼", "工項名稱", "單位", "計算式", "計算結果"]
        self._auto_fit_quantity_columns()

        corner = self._label(self.quantity_parent, "", "header")
        corner.grid(row=0, column=0, sticky="nsew")
        for col in range(QTY_COLUMNS):
            label = self._label(self.quantity_parent, self._excel_column_name(col + 1), "header")
            label.grid(row=0, column=col + 1, sticky="nsew")
            self.quantity_parent.columnconfigure(col + 1, weight=1, minsize=self.quantity_col_widths[col])
            label.bind("<Enter>", lambda event: event.widget.configure(cursor="sb_h_double_arrow"))
            label.bind("<Leave>", lambda event: event.widget.configure(cursor=""))
            label.bind("<ButtonPress-1>", lambda event, c=col: self._start_quantity_resize(event, c))
            label.bind("<B1-Motion>", lambda event, c=col: self._drag_quantity_resize(event, c))

        header_row_label = self._label(self.quantity_parent, "1", "row_header")
        header_row_label.grid(row=1, column=0, sticky="nsew")
        for col, text in enumerate(headers):
            label = self._label(self.quantity_parent, text, "field_header")
            label.grid(row=1, column=col + 1, sticky="nsew")
            label.bind("<Enter>", lambda event: event.widget.configure(cursor="sb_h_double_arrow"))
            label.bind("<Leave>", lambda event: event.widget.configure(cursor=""))
            label.bind("<ButtonPress-1>", lambda event, c=col: self._start_quantity_resize(event, c))
            label.bind("<B1-Motion>", lambda event, c=col: self._drag_quantity_resize(event, c))

        for row_index, row_data in enumerate(self.quantity_data):
            visual_row = row_index + 2
            row_label = self._label(self.quantity_parent, str(visual_row), "row_header")
            row_label.grid(row=visual_row, column=0, sticky="nsew")

            row_vars = []
            row_entries = []
            self.quantity_vars.append(row_vars)
            self.quantity_entries.append(row_entries)
            for col in range(QTY_COLUMNS):
                value = tk.StringVar(value=row_data[col])
                row_vars.append(value)

                is_readonly = col not in QTY_EDITABLE_COLUMNS
                entry = tk.Entry(
                    self.quantity_parent,
                    textvariable=value,
                    width=max(8, self.quantity_col_widths[col] // 10),
                    relief="solid",
                    borderwidth=1,
                    bg="#f3f4f6" if is_readonly else "#ffffff",
                    readonlybackground="#f3f4f6",
                    justify="center" if col in (0, QTY_RESULT_INDEX) else "left",
                    state="readonly" if is_readonly else "normal",
                )
                entry.grid(row=visual_row, column=col + 1, sticky="nsew", padx=1, pady=1)
                entry.bind("<FocusIn>", lambda _event, r=row_index, c=col: self._remember_quantity_focus(r, c))
                entry.bind("<FocusOut>", lambda _event: self._commit_edit_history())
                entry.bind("<Return>", lambda event, r=row_index, c=col: self._handle_quantity_enter(event, r, c))
                entry.bind("<Up>", lambda event, r=row_index, c=col: self._handle_quantity_arrow(event, r, c, -1, 0))
                entry.bind("<Down>", lambda event, r=row_index, c=col: self._handle_quantity_arrow(event, r, c, 1, 0))
                entry.bind("<Left>", lambda event, r=row_index, c=col: self._handle_quantity_arrow(event, r, c, 0, -1))
                entry.bind("<Right>", lambda event, r=row_index, c=col: self._handle_quantity_arrow(event, r, c, 0, 1))
                row_entries.append(entry)

                if col == QTY_EXPR_INDEX:
                    value.trace_add("write", lambda *_args, r=row_index: self._update_quantity_expression(r))
                elif col in QTY_SYNC_COLUMNS:
                    value.trace_add("write", lambda *_args, r=row_index, c=col: self._update_quantity_synced_cell(r, c))

    def _build_table(self):
        parent = self.table_parent
        self._is_rendering = True
        for child in parent.winfo_children():
            child.destroy()

        self.cell_vars = []
        self.cell_entries = []
        self.level_labels = []
        self.code_entries = []
        self.code_status_labels = []
        self.row_widgets = []

        widths = [10, 16, 22, 18, 10, 12, 12, 12, 24, 12]
        headers = ["項次", "工項號碼", "工項名稱", "規格", "單位", "數量", "單價", "複價", "備註", ""]

        corner = self._label(parent, "", "header")
        corner.grid(row=0, column=0, sticky="nsew")

        level_column = self._label(parent, "層級", "header")
        level_column.grid(row=0, column=1, sticky="nsew")
        parent.columnconfigure(1, weight=0, minsize=70)

        for col in range(DATA_COLUMNS):
            label = self._label(parent, self._excel_column_name(col + 1), "header")
            label.grid(row=0, column=col + 2, sticky="nsew")
            parent.columnconfigure(col + 2, weight=1, minsize=widths[col] * 10)

        header_row_label = self._label(parent, "1", "row_header")
        header_row_label.grid(row=1, column=0, sticky="nsew")

        level_header = self._label(parent, "層級", "field_header")
        level_header.grid(row=1, column=1, sticky="nsew")

        for col, text in enumerate(headers):
            label = self._label(parent, text, "field_header")
            label.grid(row=1, column=col + 2, sticky="nsew")

        for row in range(len(self.rows_data)):
            visual_row = row + 2
            row_label = self._label(parent, str(visual_row), "row_header")
            row_label.grid(row=visual_row, column=0, sticky="nsew")
            row_widget_group = [row_label]

            level_label = tk.Label(
                parent,
                text="",
                width=6,
                bg="#ffffff",
                relief="solid",
                borderwidth=1,
                padx=6,
                pady=5,
            )
            level_label.grid(row=visual_row, column=1, sticky="nsew", padx=1, pady=1)
            self.level_labels.append(level_label)
            row_widget_group.append(level_label)

            row_vars = []
            row_entries = []
            self.cell_vars.append(row_vars)
            self.cell_entries.append(row_entries)
            for col in range(DATA_COLUMNS):
                value = tk.StringVar(value=self.rows_data[row][col])
                row_vars.append(value)

                if col == ITEM_COLUMN_INDEX:
                    entry = tk.Entry(
                        parent,
                        textvariable=value,
                        width=widths[col],
                        relief="solid",
                        borderwidth=1,
                        bg="#f3f4f6",
                        readonlybackground="#f3f4f6",
                        justify="center",
                        state="readonly",
                    )
                elif col == CODE_COLUMN_INDEX:
                    entry = tk.Entry(
                        parent,
                        textvariable=value,
                        width=widths[col],
                        relief="solid",
                        borderwidth=1,
                        bg="#ffffff",
                    )
                    self.code_entries.append(entry)
                else:
                    entry = ttk.Entry(parent, textvariable=value, width=widths[col])
                entry.grid(row=visual_row, column=col + 2, sticky="nsew", padx=1, pady=1)
                entry.bind("<FocusIn>", lambda _event, r=row, c=col: self._remember_focus(r, c))
                entry.bind("<FocusOut>", lambda _event: self._commit_edit_history())
                entry.bind("<Double-Button-1>", lambda event, r=row, c=col: self._lock_cell_edit(event, r, c))
                entry.bind("<Return>", lambda event, r=row, c=col: self._handle_enter(event, r, c))
                entry.bind("<Up>", lambda event, r=row, c=col: self._handle_arrow(event, r, c, -1, 0))
                entry.bind("<Down>", lambda event, r=row, c=col: self._handle_arrow(event, r, c, 1, 0))
                entry.bind("<Left>", lambda event, r=row, c=col: self._handle_arrow(event, r, c, 0, -1))
                entry.bind("<Right>", lambda event, r=row, c=col: self._handle_arrow(event, r, c, 0, 1))
                row_entries.append(entry)
                row_widget_group.append(entry)

                if col != ITEM_COLUMN_INDEX:
                    value.trace_add("write", lambda *_args, r=row, c=col: self._update_cell_data(r, c))

                if col == CODE_COLUMN_INDEX:
                    status = tk.Label(parent, text="", width=2, bg="#ffffff")
                    status.grid(row=visual_row, column=col + 2, sticky="e", padx=(0, 3))
                    self.code_status_labels.append(status)
                    row_widget_group.append(status)
                    value.trace_add("write", lambda *_args, index=row: self.validate_all(index))

            self.row_widgets.append(row_widget_group)

        self._configure_styles()
        self._is_rendering = False
        self.validate_all(0)
        self._apply_visibility()
        if self.quantity_parent is not None:
            self._build_quantity_table()

    def _configure_styles(self):
        style = ttk.Style(self)
        style.configure("TEntry", padding=4)
        style.configure("TCombobox", padding=4)

    def _label(self, parent, text, kind):
        colors = {
            "header": ("#e7e9ed", "#202124"),
            "row_header": ("#f2f3f5", "#333333"),
            "field_header": ("#dfe6f3", "#111827"),
        }
        bg, fg = colors[kind]
        label = tk.Label(
            parent,
            text=text,
            bg=bg,
            fg=fg,
            relief="solid",
            borderwidth=1,
            padx=6,
            pady=6,
            font=("Arial", 11, "bold" if kind != "row_header" else "normal"),
        )
        return label

    def _grid_label(self, parent, text, bg, bold=False, width=12, anchor="center"):
        label = tk.Label(
            parent,
            text=text,
            bg=bg,
            fg="#111827",
            relief="solid",
            borderwidth=1,
            padx=6,
            pady=6,
            width=width,
            anchor=anchor,
            justify="left",
            wraplength=max(width * 8, 80),
            font=("Arial", 11, "bold" if bold else "normal"),
        )
        return label

    def _template_cell_label(self, parent, cell):
        value = "" if cell.value is None else str(cell.value)
        bg = self._cell_fill_color(cell)
        fg = self._cell_font_color(cell)
        anchor = self._cell_anchor(cell)
        justify = "center" if anchor == "center" else "left"
        size = int(cell.font.sz) if cell.font and cell.font.sz else 10
        weight = "bold" if cell.font and cell.font.bold else "normal"

        return tk.Label(
            parent,
            text=value,
            bg=bg,
            fg=fg,
            relief="solid",
            borderwidth=1,
            padx=4,
            pady=3,
            anchor=anchor,
            justify=justify,
            wraplength=220,
            font=("Arial", max(8, min(size, 18)), weight),
        )

    def _cell_fill_color(self, cell):
        fill = cell.fill
        if not fill or fill.fill_type != "solid":
            return "#ffffff"
        return self._excel_color_to_hex(fill.fgColor, "#ffffff")

    def _cell_font_color(self, cell):
        if not cell.font or not cell.font.color:
            return "#111827"
        return self._excel_color_to_hex(cell.font.color, "#111827")

    def _excel_color_to_hex(self, color, default):
        if color is None or color.type != "rgb":
            return default

        rgb = color.rgb
        if rgb is None:
            return default

        rgb_text = str(rgb)
        if len(rgb_text) == 8:
            return f"#{rgb_text[-6:]}"
        if len(rgb_text) == 6:
            return f"#{rgb_text}"
        return default

    def _cell_anchor(self, cell):
        horizontal = cell.alignment.horizontal if cell.alignment else None
        vertical = cell.alignment.vertical if cell.alignment else None
        if horizontal == "center":
            return "center"
        if horizontal == "right":
            return "e"
        if vertical == "top":
            return "nw"
        return "w"

    def validate_row(self, row_index):
        code_var = self.cell_vars[row_index][CODE_COLUMN_INDEX]
        original = code_var.get()
        normalized = self._normalize_code(original)

        if original != normalized:
            code_var.set(normalized)
            return

        if not normalized:
            self.level_labels[row_index].configure(text="")
            if self._row_has_detail_data(row_index):
                self._set_row_status(row_index, "!", "#fee2e2", "本列已有資料，B 欄必須輸入工項號碼。")
            else:
                self._set_row_status(row_index, "", "#ffffff", "請輸入工項號碼。")
            return

        if self._has_blank_code_before(row_index):
            self._set_row_status(row_index, "!", "#fee2e2", "B 欄不能跳過上方空白列輸入工項號碼。")
            self.level_labels[row_index].configure(text=str(min(len(normalized), 8)))
            return

        if self._code_exists_elsewhere(row_index, normalized):
            self._set_row_status(row_index, "!", "#fee2e2", f"B 欄工項號碼 {normalized} 已經存在。")
            self.level_labels[row_index].configure(text=str(min(len(normalized), 8)))
            return

        codes = self._valid_existing_codes(self._previous_raw_codes(row_index))
        level = len(normalized)
        self.level_labels[row_index].configure(text=str(level) if level <= 8 else "8")
        is_valid, message = self._validate_code(normalized, level, codes)
        if is_valid:
            is_valid, message = self._validate_against_previous_row(row_index, normalized)

        if is_valid:
            self._set_row_status(row_index, "OK", "#d1fae5", message)
        else:
            self._set_row_status(row_index, "!", "#fee2e2", message)

    def validate_all(self, active_row_index):
        if self._is_validating or self._is_rendering:
            return

        self._is_validating = True
        try:
            self._sync_model_from_vars()
            for row_index in range(len(self.cell_vars)):
                self.validate_row(row_index)
            if 0 <= active_row_index < len(self.cell_vars):
                self.validate_row(active_row_index)
            self._renumber_items()
            self._apply_visibility()
        finally:
            self._is_validating = False

    def _previous_raw_codes(self, row_index):
        codes = []
        for row_vars in self.cell_vars[:row_index]:
            code = self._normalize_code(row_vars[CODE_COLUMN_INDEX].get())
            if code:
                codes.append(code)
        return codes

    def _has_blank_code_before(self, row_index):
        for row_vars in self.cell_vars[:row_index]:
            if not self._normalize_code(row_vars[CODE_COLUMN_INDEX].get()):
                return True
        return False

    def _code_exists_elsewhere(self, row_index, code):
        for other_row, row_vars in enumerate(self.cell_vars):
            if other_row == row_index:
                continue
            other_code = self._normalize_code(row_vars[CODE_COLUMN_INDEX].get())
            if other_code == code:
                return True
        return False

    def _row_has_detail_data(self, row_index):
        row_vars = self.cell_vars[row_index]
        for column_index, value in enumerate(row_vars):
            if column_index in (ITEM_COLUMN_INDEX, CODE_COLUMN_INDEX):
                continue
            if value.get().strip():
                return True
        return False

    def _valid_existing_codes(self, raw_codes):
        valid_codes = set()
        for code in raw_codes:
            level = len(code)
            is_valid, _message = self._validate_code(code, level, valid_codes)
            if is_valid:
                valid_codes.add(code)
        return valid_codes

    def _normalize_code(self, code):
        normalized = code.strip().upper()
        if normalized.startswith("O"):
            normalized = "0" + normalized[1:]
        return normalized

    def _validate_code(self, code, level, existing_codes):
        if level > 8:
            return False, "工項號碼最多只能編到第 8 層。"

        if len(code) != level:
            return False, f"第 {level} 層的工項號碼長度必須為 {level} 碼。"

        if level == 1:
            if code != "0":
                return False, "第 1 層只能輸入 0。"
            return True, "第 1 層工項號碼正確。"

        if not CODE_PATTERN.fullmatch(code):
            return False, "工項號碼必須以 0 開頭，後續只能使用 1~9、A~Z。"

        parent = code[:-1]
        if parent not in existing_codes:
            return False, f"必須先建立上一層 {parent}。"

        current_char = code[-1]
        if current_char == CODE_ORDER[0]:
            return True, "工項號碼正確。"

        previous_char = CODE_ORDER[CODE_ORDER.index(current_char) - 1]
        previous_code = f"{parent}{previous_char}"
        if parent == "0" and current_char.isalpha():
            return True, "工項號碼正確。"

        if previous_code not in existing_codes:
            return False, f"必須先建立前一個同層編碼 {previous_code}。"

        return True, "工項號碼正確。"

    def _validate_against_previous_row(self, row_index, code):
        previous_code = self._previous_code(row_index)
        if not previous_code:
            return True, "工項號碼正確。"

        allowed_codes = self._allowed_next_codes(previous_code)
        if code in allowed_codes:
            return True, "工項號碼正確。"

        display = "、".join(allowed_codes)
        return False, f"必須接續上一行 {previous_code}，下一個可用編碼為 {display}。"

    def _previous_code(self, row_index):
        for row_vars in reversed(self.cell_vars[:row_index]):
            code = self._normalize_code(row_vars[CODE_COLUMN_INDEX].get())
            if code:
                return code
        return ""

    def _allowed_next_codes(self, previous_code):
        allowed_codes = []
        if len(previous_code) < 8:
            allowed_codes.append(previous_code + CODE_ORDER[0])

        if previous_code == "0":
            allowed_codes.extend(f"0{letter}" for letter in "ABCDEFGHIJKLMNOPQRSTUVWXYZ")

        sibling = self._next_sibling_code(previous_code)
        if sibling:
            allowed_codes.append(sibling)

        if len(previous_code) == 2 and previous_code.startswith("0") and previous_code[-1] in "123456789":
            allowed_codes.extend(f"0{letter}" for letter in "ABCDEFGHIJKLMNOPQRSTUVWXYZ")

        for level in range(len(previous_code) - 1, 1, -1):
            ancestor = previous_code[:level]
            sibling = self._next_sibling_code(ancestor)
            if sibling:
                allowed_codes.append(sibling)
            if ancestor.startswith("0") and len(ancestor) == 2 and ancestor[-1] in "123456789":
                allowed_codes.extend(f"0{letter}" for letter in "ABCDEFGHIJKLMNOPQRSTUVWXYZ")

        seen = set()
        unique_codes = []
        for code in allowed_codes:
            if code not in seen:
                unique_codes.append(code)
                seen.add(code)
        return unique_codes

    def _next_sibling_code(self, code):
        if len(code) <= 1:
            return ""

        last_char = code[-1]
        if last_char not in CODE_ORDER:
            return ""

        index = CODE_ORDER.index(last_char)
        if index + 1 >= len(CODE_ORDER):
            return ""

        return code[:-1] + CODE_ORDER[index + 1]

    def _set_row_status(self, row_index, text, color, message):
        label = self.code_status_labels[row_index]
        label.configure(text=text, bg=color)
        self.code_entries[row_index].configure(bg=color)
        self.status_var.set(f"第 {row_index + 2} 列：{message}")

    def _remember_focus(self, row_index, column_index):
        self.focused_cell = (row_index, column_index)
        self._begin_edit_history()

    def _update_cell_data(self, row_index, column_index):
        if self._is_rendering:
            return
        value = self.cell_vars[row_index][column_index].get()
        self.rows_data[row_index][column_index] = value
        quantity_column = self._source_to_quantity_column(column_index)
        if quantity_column is not None and row_index < len(self.quantity_data):
            self.quantity_data[row_index][quantity_column] = value
            if row_index < len(self.quantity_vars) and quantity_column < len(self.quantity_vars[row_index]):
                target_var = self.quantity_vars[row_index][quantity_column]
                if target_var.get() != value:
                    target_var.set(value)

    def _sync_model_from_vars(self):
        for row_index, row_vars in enumerate(self.cell_vars):
            for column_index, value in enumerate(row_vars):
                self.rows_data[row_index][column_index] = value.get()

    def _sync_quantity_source(self):
        old_expressions = []
        for row in self.quantity_data:
            old_expressions.append(row[QTY_EXPR_INDEX] if len(row) > QTY_EXPR_INDEX else "")

        quantity_rows = []
        for row_index, source_row in enumerate(self.rows_data):
            row = []
            for source_column in QTY_SOURCE_COLUMNS[:QTY_EXPR_INDEX]:
                row.append(source_row[source_column] if source_column is not None and source_column < len(source_row) else "")
            expression = old_expressions[row_index] if row_index < len(old_expressions) else ""
            result = source_row[QTY_RESULT_INDEX] if len(source_row) > QTY_RESULT_INDEX else ""
            row.extend([expression, result])
            quantity_rows.append(row)

        while len(quantity_rows) < MIN_DATA_ROWS:
            quantity_rows.append(["" for _column in range(QTY_COLUMNS)])
        self.quantity_data = quantity_rows

    def _sync_quantity_from_vars(self):
        for row_index, row_vars in enumerate(self.quantity_vars):
            if row_index >= len(self.quantity_data):
                continue
            for column_index in QTY_SYNC_COLUMNS:
                self.quantity_data[row_index][column_index] = row_vars[column_index].get()
                source_column = self._quantity_to_source_column(column_index)
                if source_column is not None and row_index < len(self.rows_data):
                    self.rows_data[row_index][source_column] = row_vars[column_index].get()
            self.quantity_data[row_index][QTY_EXPR_INDEX] = row_vars[QTY_EXPR_INDEX].get()
            self.quantity_data[row_index][QTY_RESULT_INDEX] = row_vars[QTY_RESULT_INDEX].get()

    def _update_quantity_expression(self, row_index):
        if row_index >= len(self.quantity_data):
            return
        self.quantity_data[row_index][QTY_EXPR_INDEX] = self.quantity_vars[row_index][QTY_EXPR_INDEX].get()

    def _update_quantity_synced_cell(self, row_index, column_index):
        if row_index >= len(self.quantity_data):
            return

        value = self.quantity_vars[row_index][column_index].get()
        self.quantity_data[row_index][column_index] = value
        source_column = self._quantity_to_source_column(column_index)
        if source_column is not None and row_index < len(self.rows_data):
            self.rows_data[row_index][source_column] = value

        if source_column is not None and row_index < len(self.cell_vars) and source_column < len(self.cell_vars[row_index]):
            target_var = self.cell_vars[row_index][source_column]
            if target_var.get() != value:
                target_var.set(value)

    def _quantity_to_source_column(self, quantity_column):
        if 0 <= quantity_column < len(QTY_SOURCE_COLUMNS):
            return QTY_SOURCE_COLUMNS[quantity_column]
        return None

    def _source_to_quantity_column(self, source_column):
        for quantity_column, mapped_source in enumerate(QTY_SOURCE_COLUMNS):
            if mapped_source == source_column:
                return quantity_column
        return None

    def _auto_fit_quantity_columns(self):
        for column_index in (0, 1, QTY_RESULT_INDEX):
            max_length = 4
            header_lengths = {0: 4, 1: 8, QTY_RESULT_INDEX: 8}
            max_length = header_lengths[column_index]
            for row in self.quantity_data:
                if column_index < len(row):
                    max_length = max(max_length, len(str(row[column_index])))
            self.quantity_col_widths[column_index] = max(70, min(260, max_length * 12 + 28))

    def _start_quantity_resize(self, event, column_index):
        self.quantity_resize = (column_index, event.x_root, self.quantity_col_widths[column_index])

    def _drag_quantity_resize(self, event, column_index):
        if not self.quantity_resize:
            return

        resize_column, start_x, start_width = self.quantity_resize
        if resize_column != column_index:
            return

        new_width = max(60, start_width + event.x_root - start_x)
        self.quantity_col_widths[column_index] = new_width
        self.quantity_parent.columnconfigure(column_index + 1, minsize=new_width)

    def _snapshot_state(self):
        return {
            "rows": [list(row) for row in self.rows_data],
            "quantity_rows": [list(row) for row in self.quantity_data],
            "project": {
                "name": self.project_name_var.get(),
                "execution_no": self.execution_no_var.get(),
                "budget_year": self.budget_year_var.get(),
                "location": self.project_location_var.get(),
                "content": self.project_content_text.get("1.0", "end-1c"),
            },
        }

    def _restore_snapshot(self, snapshot):
        self._history_suspended = True
        try:
            project = snapshot["project"]
            self.project_name_var.set(project["name"])
            self.execution_no_var.set(project["execution_no"])
            self.budget_year_var.set(project["budget_year"])
            self.project_location_var.set(project["location"])
            self.project_content_text.delete("1.0", tk.END)
            self.project_content_text.insert("1.0", project["content"])
            self.rows_data = [list(row) for row in snapshot["rows"]]
            self.quantity_data = [list(row) for row in snapshot["quantity_rows"]]
            self._build_table()
        finally:
            self._history_suspended = False

    def _begin_edit_history(self):
        if self._history_suspended:
            return
        self._edit_snapshot = self._snapshot_state()

    def _commit_edit_history(self):
        if self._history_suspended or self._edit_snapshot is None:
            return
        current = self._snapshot_state()
        if current != self._edit_snapshot:
            self.undo_stack.append(self._edit_snapshot)
            self.undo_stack = self.undo_stack[-20:]
            self.redo_stack.clear()
        self._edit_snapshot = None

    def _record_history(self):
        if self._history_suspended:
            return
        self.undo_stack.append(self._snapshot_state())
        self.undo_stack = self.undo_stack[-20:]
        self.redo_stack.clear()

    def undo_last_change(self):
        if not self.undo_stack:
            self.top_message_var.set("沒有可回復的歷史操作。")
            return
        self.redo_stack.append(self._snapshot_state())
        self.redo_stack = self.redo_stack[-20:]
        self._restore_snapshot(self.undo_stack.pop())
        self.top_message_var.set("已回復上一個歷史操作。")

    def redo_last_change(self):
        if not self.redo_stack:
            self.top_message_var.set("沒有可往後的歷史操作。")
            return
        self.undo_stack.append(self._snapshot_state())
        self.undo_stack = self.undo_stack[-20:]
        self._restore_snapshot(self.redo_stack.pop())
        self.top_message_var.set("已重做下一個歷史操作。")

    def _remember_quantity_focus(self, row_index, column_index):
        self.focused_cell = (row_index, min(column_index, DATA_COLUMNS - 1))
        self._begin_edit_history()

    def _handle_quantity_enter(self, event, row_index, column_index):
        self._commit_edit_history()
        next_row = min(row_index + 1, len(self.quantity_entries) - 1)
        self._focus_quantity_cell(next_row, column_index)
        return "break"

    def _handle_quantity_arrow(self, event, row_index, column_index, row_delta, column_delta):
        target_row = max(0, min(row_index + row_delta, len(self.quantity_entries) - 1))
        target_column = max(0, min(column_index + column_delta, QTY_COLUMNS - 1))
        self._focus_quantity_cell(target_row, target_column)
        return "break"

    def _focus_quantity_cell(self, row_index, column_index):
        if row_index >= len(self.quantity_entries):
            return
        entry = self.quantity_entries[row_index][column_index]
        entry.focus_set()
        if hasattr(entry, "icursor"):
            entry.icursor(tk.END)

    def calculate_quantity_page(self):
        self._sync_model_from_vars()
        self._sync_quantity_from_vars()
        self._sync_quantity_source()
        self._record_history()
        errors = []

        for row_index, row in enumerate(self.quantity_data):
            expression = row[QTY_EXPR_INDEX].strip()
            result = ""
            if expression:
                try:
                    result = self._format_number(self._safe_eval_expression(expression))
                except ValueError as error:
                    result = "錯誤"
                    errors.append(f"第 {row_index + 2} 列：{error}")
            row[QTY_RESULT_INDEX] = result
            if row_index < len(self.rows_data):
                self.rows_data[row_index][6] = result
            if row_index < len(self.cell_vars):
                target_var = self.cell_vars[row_index][6]
                if target_var.get() != result:
                    target_var.set(result)

        self._build_quantity_table()
        if errors:
            messagebox.showwarning("第3分頁計算錯誤", "\n".join(errors[:20]))
            self.top_message_var.set(f"第3分頁計算完成，但有 {len(errors)} 筆錯誤。")
        else:
            self.top_message_var.set("第3分頁計算完成。")

    def _safe_eval_expression(self, expression):
        try:
            tree = ast.parse(expression, mode="eval")
        except SyntaxError as error:
            raise ValueError(f"算式語法錯誤：{error.msg}") from error
        return self._eval_ast_node(tree.body)

    def _eval_ast_node(self, node):
        if isinstance(node, ast.Constant) and isinstance(node.value, (int, float)):
            return node.value
        if isinstance(node, ast.BinOp) and type(node.op) in QTY_OPERATORS:
            left = self._eval_ast_node(node.left)
            right = self._eval_ast_node(node.right)
            if isinstance(node.op, ast.Div) and right == 0:
                raise ValueError("除數不能為 0")
            return QTY_OPERATORS[type(node.op)](left, right)
        if isinstance(node, ast.UnaryOp) and type(node.op) in QTY_OPERATORS:
            operand = self._eval_ast_node(node.operand)
            return QTY_OPERATORS[type(node.op)](operand)
        raise ValueError("只允許數字與 +、-、*、/、括號")

    def _format_number(self, value):
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(round(value, 8))

    def _renumber_items(self):
        number = 1
        for row_index, row_vars in enumerate(self.cell_vars):
            item_var = row_vars[ITEM_COLUMN_INDEX]
            code = self._normalize_code(row_vars[CODE_COLUMN_INDEX].get())
            new_item = str(number) if code else ""
            if item_var.get() != new_item:
                item_entry = self.cell_entries[row_index][ITEM_COLUMN_INDEX]
                item_entry.configure(state="normal")
                item_var.set(new_item)
                item_entry.configure(state="readonly")
            self.rows_data[row_index][ITEM_COLUMN_INDEX] = new_item
            if row_index < len(self.quantity_data):
                self.quantity_data[row_index][ITEM_COLUMN_INDEX] = new_item
            if row_index < len(self.quantity_vars):
                quantity_item_var = self.quantity_vars[row_index][ITEM_COLUMN_INDEX]
                if quantity_item_var.get() != new_item:
                    quantity_item_var.set(new_item)
            if code:
                number += 1

    def _row_has_code_error(self, row_index):
        if row_index < 0 or row_index >= len(self.code_entries):
            return False
        return self.code_entries[row_index].cget("bg") == "#fee2e2"

    def _handle_enter(self, event, row_index, column_index):
        self._commit_edit_history()
        self.editing_cell = None
        self.validate_all(row_index)
        if column_index >= CODE_COLUMN_INDEX and self._row_has_code_error(row_index):
            self.status_var.set(f"第 {row_index + 2} 列 B 欄檢查有錯誤，修正後才會跳到下一行。")
            return "break"

        next_row = row_index + 1
        if next_row >= len(self.rows_data):
            self.rows_data.append(["" for _column in range(DATA_COLUMNS)])
            self._build_table()

        self.after(1, lambda: self._focus_cell(next_row, column_index))
        return "break"

    def _handle_arrow(self, event, row_index, column_index, row_delta, column_delta):
        if self.editing_cell == (row_index, column_index):
            return None

        target_column = max(0, min(column_index + column_delta, DATA_COLUMNS - 1))
        target_row = row_index

        if row_delta:
            target_row = self._next_visible_row(row_index, row_delta)

        self._focus_cell(target_row, target_column)
        return "break"

    def _lock_cell_edit(self, event, row_index, column_index):
        self.editing_cell = (row_index, column_index)
        return None

    def _next_visible_row(self, row_index, row_delta):
        target_row = row_index + row_delta
        while 0 <= target_row < len(self.rows_data):
            if self._row_is_visible(target_row):
                return target_row
            target_row += row_delta
        return row_index

    def _focus_cell(self, row_index, column_index):
        if row_index >= len(self.cell_entries):
            return
        self.editing_cell = None
        entry = self.cell_entries[row_index][column_index]
        entry.focus_set()
        if hasattr(entry, "icursor"):
            entry.icursor(tk.END)

    def _scroll_to_row(self, row_index):
        if self.table_canvas is None or not self.rows_data:
            return

        self.update_idletasks()
        fraction = max(0, min(row_index / max(len(self.rows_data), 1), 1))
        self.table_canvas.yview_moveto(fraction)

    def insert_rows_at_cursor(self):
        self._sync_model_from_vars()
        try:
            count = int(self.insert_count_var.get())
        except ValueError:
            count = 1

        count = max(1, min(count, 100))
        row_index, column_index = self.focused_cell
        self._record_history()
        insert_index = max(0, min(row_index + 1, len(self.rows_data)))
        blank_rows = [["" for _column in range(DATA_COLUMNS)] for _row in range(count)]
        self.rows_data[insert_index:insert_index] = blank_rows
        self.top_message_var.set(f"已從第 {insert_index + 2} 列插入 {count} 行。")
        self._build_table()
        self.after(1, lambda: self._focus_cell(insert_index, min(max(column_index, CODE_COLUMN_INDEX), DATA_COLUMNS - 1)))

    def delete_row_at_cursor(self):
        self._sync_model_from_vars()
        row_index, column_index = self.focused_cell
        if not self.rows_data:
            return

        row_index = max(0, min(row_index, len(self.rows_data) - 1))
        self._record_history()
        deleted_row = list(self.rows_data[row_index])
        del self.rows_data[row_index]
        self.deleted_rows_stack.append((row_index, deleted_row))
        if len(self.deleted_rows_stack) > 20:
            self.deleted_rows_stack.pop(0)
        while len(self.rows_data) < MIN_DATA_ROWS:
            self.rows_data.append(["" for _column in range(DATA_COLUMNS)])

        self.top_message_var.set(f"已刪除第 {row_index + 2} 列。")
        self._build_table()
        focus_row = min(row_index, len(self.rows_data) - 1)
        self.after(1, lambda: self._focus_cell(focus_row, min(max(column_index, CODE_COLUMN_INDEX), DATA_COLUMNS - 1)))

    def undo_delete_row(self):
        if not self.deleted_rows_stack:
            self.top_message_var.set("目前沒有可復原的刪除記錄。")
            return

        self._sync_model_from_vars()
        row_index, row_data = self.deleted_rows_stack.pop()
        self._record_history()
        row_index = max(0, min(row_index, len(self.rows_data)))
        self.rows_data[row_index:row_index] = [row_data]
        self.top_message_var.set(f"已復原刪除第 {row_index + 2} 列。")
        self._build_table()
        self.after(1, lambda: self._focus_cell(row_index, CODE_COLUMN_INDEX))

    def jump_to_code(self):
        code = self._normalize_code(self.jump_code_var.get())
        if not code:
            self.top_message_var.set("請輸入要跳到的工項編碼。")
            return

        self._sync_model_from_vars()
        for row_index, row in enumerate(self.rows_data):
            if self._normalize_code(row[CODE_COLUMN_INDEX]) == code:
                self.show_all_rows()
                self._scroll_to_row(row_index)
                self._focus_cell(row_index, CODE_COLUMN_INDEX)
                self.top_message_var.set(f"已跳到工項 {code}。")
                return

        self.top_message_var.set(f"找不到工項 {code}。")

    def search_keyword(self):
        keyword = self.search_keyword_var.get().strip().lower()
        if not keyword:
            self.top_message_var.set("請輸入搜尋關鍵字。")
            return

        self._sync_model_from_vars()
        start_row, start_column = self.focused_cell
        positions = []
        for row_index, row in enumerate(self.rows_data):
            for column_index, value in enumerate(row):
                if keyword in str(value).lower():
                    positions.append((row_index, column_index))

        if not positions:
            self.top_message_var.set(f"找不到關鍵字：{self.search_keyword_var.get().strip()}")
            return

        target = positions[0]
        for position in positions:
            if position > (start_row, start_column):
                target = position
                break

        self.show_all_rows()
        self._scroll_to_row(target[0])
        self._focus_cell(target[0], target[1])
        self.top_message_var.set(f"找到關鍵字：第 {target[0] + 2} 列。")

    def save_record(self):
        self.validate_all(self.focused_cell[0])
        payload = self._record_payload()
        path = self._save_path_from_input()

        try:
            path.parent.mkdir(parents=True, exist_ok=True)
            path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        except OSError as error:
            messagebox.showerror("儲存失敗", str(error))
            return

        self.save_filename_var.set(path.name)
        self.top_message_var.set(f"已儲存記錄：{path.name}")

    def open_record(self):
        filename = self.save_filename_var.get().strip()
        path = self._path_from_filename(filename) if filename else None
        if path is None or not path.exists():
            selected = filedialog.askopenfilename(
                title="開啟記錄",
                filetypes=[("JSON 記錄檔", "*.json"), ("所有檔案", "*.*")],
            )
            if not selected:
                return
            path = Path(selected)

        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError) as error:
            messagebox.showerror("開啟失敗", str(error))
            return

        self._load_record_payload(payload)
        self.save_filename_var.set(path.name)
        self.deleted_rows_stack = []
        self.top_message_var.set(f"已開啟記錄：{path.name}")

    def export_cover_content_excel(self):
        try:
            from openpyxl import Workbook, load_workbook
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        except ImportError as error:
            messagebox.showerror("輸出失敗", f"缺少 openpyxl 套件：{error}")
            return

        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        default_name = f"預算書封面內容_{timestamp}.xlsx"
        output = filedialog.asksaveasfilename(
            title="輸出第2分頁",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel 活頁簿", "*.xlsx"), ("所有檔案", "*.*")],
        )
        if not output:
            return

        path = Path(output)
        if DEFAULT_COVER_TEMPLATE.exists():
            try:
                workbook = load_workbook(DEFAULT_COVER_TEMPLATE)
                self._apply_cover_print_settings(workbook.active, Border, Side)
                workbook.save(path)
            except OSError as error:
                messagebox.showerror("輸出失敗", str(error))
                return

            self.top_message_var.set(f"已依範本格式輸出第2分頁：{path.name}")
            return

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "預算書封面內容"
        sheet.freeze_panes = "A4"

        title_fill = PatternFill("solid", fgColor="1F4E78")
        header_fill = PatternFill("solid", fgColor="D9EAF7")
        section_fill = PatternFill("solid", fgColor="E2F0D9")
        thin_side = Side(style="thin", color="808080")
        border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        sheet.merge_cells("A1:F1")
        sheet["A1"] = "預算書封面內容"
        sheet["A1"].font = Font(bold=True, color="FFFFFF", size=16)
        sheet["A1"].fill = title_fill
        sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
        sheet.row_dimensions[1].height = 28

        sheet.merge_cells("A2:F2")
        sheet["A2"] = "資料來源與對應欄位先保留待設定，本頁只顯示封面內容項目。"
        sheet["A2"].font = Font(color="666666")
        sheet["A2"].alignment = Alignment(horizontal="left", vertical="center")

        headers = ["類別", "欄位代號", "封面內容", "封面位置", "資料來源", "對應欄位"]
        for column_index, header in enumerate(headers, start=1):
            cell = sheet.cell(row=3, column=column_index, value=header)
            cell.fill = header_fill
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        for row_index, row_values in enumerate(self._cover_content_rows(), start=4):
            for column_index, value in enumerate(row_values, start=1):
                cell = sheet.cell(row=row_index, column=column_index, value=value)
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                if column_index == 1:
                    cell.fill = section_fill

        column_widths = {
            "A": 18,
            "B": 24,
            "C": 24,
            "D": 14,
            "E": 24,
            "F": 24,
        }
        for column, width in column_widths.items():
            sheet.column_dimensions[column].width = width

        self._apply_cover_print_settings(sheet, Border, Side)

        try:
            path.parent.mkdir(parents=True, exist_ok=True)
            workbook.save(path)
        except OSError as error:
            messagebox.showerror("輸出失敗", str(error))
            return

        self.top_message_var.set(f"已輸出第2分頁：{path.name}")

    def _apply_cover_print_settings(self, sheet, border_class, side_class):
        thin_side = side_class(style="thin", color="000000")
        grid_border = border_class(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        for row in sheet["A1:O21"]:
            for cell in row:
                cell.border = grid_border

        sheet.print_area = "A1:O21"

    def _record_payload(self):
        self._sync_model_from_vars()
        self._sync_quantity_from_vars()
        self._sync_quantity_source()
        return {
            "project": {
                "name": self.project_name_var.get(),
                "execution_no": self.execution_no_var.get(),
                "budget_year": self.budget_year_var.get(),
                "location": self.project_location_var.get(),
                "content": self.project_content_text.get("1.0", "end-1c"),
            },
            "rows": self.rows_data,
            "quantity_rows": self.quantity_data,
        }

    def _load_record_payload(self, payload):
        project = payload.get("project", {})
        self.project_name_var.set(project.get("name", ""))
        self.execution_no_var.set(project.get("execution_no", ""))
        self.budget_year_var.set(project.get("budget_year", ""))
        self.project_location_var.set(project.get("location", ""))
        self.project_content_text.delete("1.0", tk.END)
        self.project_content_text.insert("1.0", project.get("content", ""))

        rows = payload.get("rows", [])
        clean_rows = []
        for row in rows:
            values = list(row[:DATA_COLUMNS]) if isinstance(row, list) else []
            values.extend([""] * (DATA_COLUMNS - len(values)))
            clean_rows.append(values[:DATA_COLUMNS])

        while len(clean_rows) < MIN_DATA_ROWS:
            clean_rows.append(["" for _column in range(DATA_COLUMNS)])

        self.rows_data = clean_rows
        quantity_rows = payload.get("quantity_rows", [])
        clean_quantity_rows = []
        for row in quantity_rows:
            values = list(row[:QTY_COLUMNS]) if isinstance(row, list) else []
            values.extend([""] * (QTY_COLUMNS - len(values)))
            clean_quantity_rows.append(values[:QTY_COLUMNS])
        self.quantity_data = clean_quantity_rows
        self.show_all_rows()
        self._build_table()

    def _save_path_from_input(self):
        filename = self.save_filename_var.get().strip()
        if not filename:
            filename = "AA_T" + datetime.now().strftime("%Y%m%d%H%M%S")
        return self._path_from_filename(filename)

    def _path_from_filename(self, filename):
        path = Path(filename)
        if not path.suffix:
            path = path.with_suffix(SAVE_EXTENSION)
        if not path.is_absolute():
            path = Path.cwd() / path
        return path

    def hide_level_rows(self):
        try:
            level = int(self.hide_level_var.get())
        except ValueError:
            self.top_message_var.set("請輸入要隱藏的階層 1-8。")
            return

        self.hidden_level = max(1, min(level, 8))
        self.focus_code = ""
        self._apply_visibility()
        self.top_message_var.set(f"已隱藏第 {self.hidden_level} 層以下工項。")

    def show_only_code_tree(self):
        code = self._normalize_code(self.show_code_var.get())
        if not code:
            self.top_message_var.set("請輸入要指定顯示的工項號碼。")
            return

        self.focus_code = code
        self.hidden_level = None
        self._apply_visibility()
        self.top_message_var.set(f"只顯示 {code} 以下階層。")

    def show_all_rows(self):
        self.hidden_level = None
        self.focus_code = ""
        self.hide_level_var.set("")
        self.show_code_var.set("")
        self._apply_visibility()
        self.top_message_var.set("已恢復顯示全部。")

    def _apply_visibility(self):
        for row_index, widgets in enumerate(self.row_widgets):
            visible = self._row_is_visible(row_index)
            for widget in widgets:
                if visible:
                    widget.grid()
                else:
                    widget.grid_remove()

    def _row_is_visible(self, row_index):
        code = self._normalize_code(self.rows_data[row_index][CODE_COLUMN_INDEX])
        if self.focus_code:
            return bool(code and (code == self.focus_code or code.startswith(self.focus_code)))

        if self.hidden_level is not None:
            return not code or len(code) < self.hidden_level

        return True

    def _on_mouse_wheel(self, event):
        active_canvas = self.table_canvas
        if self.main_notebook is not None:
            try:
                current_index = self.main_notebook.index("current")
                if current_index == 1:
                    active_canvas = self.cover_canvas
                elif current_index == 2:
                    active_canvas = self.quantity_canvas
            except tk.TclError:
                active_canvas = self.table_canvas

        if active_canvas is None:
            return

        if getattr(event, "num", None) == 4:
            delta = -3
        elif getattr(event, "num", None) == 5:
            delta = 3
        else:
            if abs(event.delta) >= 120:
                delta = -1 * int(event.delta / 120)
            elif event.delta:
                delta = -1 if event.delta > 0 else 1
            else:
                delta = 0

        if delta:
            active_canvas.yview_scroll(delta, "units")

    def _excel_column_name(self, index):
        name = ""
        while index:
            index, remainder = divmod(index - 1, 26)
            name = chr(65 + remainder) + name
        return name


if __name__ == "__main__":
    app = BudgetEditor()
    app.mainloop()
